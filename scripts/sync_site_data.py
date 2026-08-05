#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de Sincronização: dados_site/*.xlsx -> React Components
Lê os arquivos Excel em dados_site/ e atualiza o conteúdo do site (Team.jsx, Partners.jsx, Papers.jsx).
"""

import os
import re
import json
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DADOS_DIR = os.path.join(BASE_DIR, "dados_site")
PAGES_DIR = os.path.join(BASE_DIR, "src", "pages")

print(f"Sincronizando dados de {DADOS_DIR}...")

# 1. SINCRONIZAR MEMBROS (membros.xlsx -> Team.jsx)
membros_file = os.path.join(DADOS_DIR, "membros.xlsx")
if os.path.exists(membros_file):
    wb = openpyxl.load_workbook(membros_file)
    ws = wb.active
    
    professors = []
    researchers = []
    former_members = []
    
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[1]: # Skip empty rows
            continue
        
        row_id, nome, cat, cargo, foto, caminho_foto, orcid, lattes, rg, email = row[:10]
        cat = str(cat or "").strip()
        nome = str(nome or "").strip()
        if nome.lower() in ["nome", "name"] or str(row_id).strip().lower() in ["id", "id_membro"]:
            continue
        cargo = str(cargo or "").strip()
        caminho_foto = str(caminho_foto or "").strip()
        orcid = str(orcid or "").strip()
        lattes = str(lattes or "").strip()
        rg = str(rg or "").strip()
        email = str(row[9] if len(row) > 9 and row[9] else "").strip()
        
        member_obj = {
            "name": nome,
            "role": cargo,
            "image": caminho_foto if caminho_foto and caminho_foto.lower() != "none" and caminho_foto != "—" else None,
            "orcid": orcid if orcid and orcid != "—" and orcid.lower() != "none" else "",
            "lattes": lattes if lattes and lattes != "—" and lattes.lower() != "none" else "",
            "researchgate": rg if rg and rg != "—" and rg.lower() != "none" else "",
            "email": email if email and email != "—" and email.lower() != "none" else ""
        }
        # Clean empty keys
        member_obj = {k: v for k, v in member_obj.items() if v}
        
        if cat == "Docente":
            professors.append(member_obj)
        elif "ex-membro" in cat.lower():
            former_obj = {"name": nome}
            if lattes and lattes != "—" and lattes.lower() != "none":
                former_obj["lattes"] = lattes
            
            # Identify subcategory
            sub_cat = "Graduação"
            if "pós" in cat.lower() or "pos" in cat.lower():
                sub_cat = "Pós-Doutorado"
            elif "doutor" in cat.lower() or "phd" in cat.lower():
                sub_cat = "Doutorado"
            elif "mestr" in cat.lower() or "msc" in cat.lower():
                sub_cat = "Mestrado"
            
            former_obj["category"] = sub_cat
            former_members.append(former_obj)
        else:
            member_obj["_category"] = cat
            researchers.append(member_obj)
            
    # Sort researchers: Pós-docs, Doutorandos, Mestrandos
    def get_cat_rank(m):
        c = str(m.get("_category", "")).lower()
        if "pós" in c or "pos" in c:
            return 1
        elif "doutor" in c:
            return 2
        elif "mestrand" in c:
            return 3
        return 4

    researchers.sort(key=lambda x: get_cat_rank(x))
    for r in researchers:
        r.pop("_category", None)

    # Sort former members: Pós-Doutores, Doutores, Mestres, IC
    def get_former_cat_rank(m):
        c = m.get("category", "")
        if c == "Pós-Doutorado":
            return 1
        elif c == "Doutorado":
            return 2
        elif c == "Mestrado":
            return 3
        return 4

    former_members.sort(key=lambda x: (get_former_cat_rank(x), x.get("name", "").lower()))
            
    print(f"Membros carregados: {len(professors)} professores, {len(researchers)} pesquisadores, {len(former_members)} ex-membros.")


# 2. SINCRONIZAR COLABORADORES E PARCEIROS (colaboradores.xlsx -> Partners.jsx)
colab_file = os.path.join(DADOS_DIR, "colaboradores.xlsx")
if os.path.exists(colab_file):
    wb = openpyxl.load_workbook(colab_file)
    ws = wb.active
    
    collaborators = []
    universities = []
    
    uni_map = {
        "uc": {"key": "uc", "name": "Universidade de Coimbra (UC)", "img": "./assets/logos/ftuc.png", "url": "https://www.uc.pt/fctuc"},
        "cura-lab": {"key": "cura-lab", "name": "CURA Lab (ADAI)", "img": "./assets/logos/curalab.png", "url": "https://cura-lab.adai.pt/"},
        "udesc": {"key": "udesc", "name": "Universidade do Estado de Santa Catarina (UDESC)", "img": "./assets/logos/udesc.png", "url": "https://www.udesc.br/cct/dau"},
        "usp": {"key": "usp", "name": "Universidade de São Paulo (USP)", "img": "./assets/logos/uspsc.png", "url": "https://eesc.usp.br/"},
        "utfpr": {"key": "utfpr", "name": "Universidade Tecnológica Federal do Paraná (UTFPR)", "img": "./assets/logos/UTFPR.png", "url": "https://www.utfpr.edu.br/"},
        "ufsc": {"key": "ufsc", "name": "Universidade Federal de Santa Catarina (UFSC)", "img": "./assets/logos/ufsc.png", "url": "https://ppgec.posgrad.ufsc.br/"},
        "ufsc-automacao": {"key": "ufsc-automacao", "name": "Universidade Federal de Santa Catarina (UFSC)", "img": "./assets/logos/ufsc.png", "url": "https://automacao.ufsc.br/"},
        "ufms": {"key": "ufms", "name": "Universidade Federal de Mato Grosso do Sul (UFMS)", "img": "./assets/logos/UFMS.png", "url": "https://faeng.ufms.br/"}
    }

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[2]:
            continue
            
        row_id = row[0]
        tipo = str(row[1] or "").strip()
        nome = str(row[2] or "").strip()
        if nome.lower() in ["nome", "name"] or str(row_id).strip().lower() in ["id", "id_colaborador"]:
            continue
        cargo = str(row[3] or "").strip()
        caminho = str(row[5] or "").strip()
        url_site = str(row[6] or "").strip()
        orcid = str(row[7] or "").strip()
        lattes = str(row[8] or "").strip()
        rg = str(row[9] or "").strip()
        scholar = str(row[10] if len(row) > 10 and row[10] else "").strip()
        areas_raw = str(row[11] if len(row) > 11 and row[11] else "").strip()
        uni_key = str(row[12] if len(row) > 12 and row[12] else "").strip()
        
        if tipo == "Colaborador":
            areas_list = [a.strip() for a in areas_raw.split(",") if a.strip()] if areas_raw else []
            uni_keys = [k.strip().lower() for k in uni_key.split(",") if k.strip()] if uni_key else []
            uni_list = [uni_map[k] for k in uni_keys if k in uni_map]
            
            colab_obj = {
                "name": nome,
                "role": cargo,
                "image": caminho if caminho and caminho.lower() != "none" else None,
                "orcid": orcid if orcid and orcid != "—" and orcid.lower() != "none" else "",
                "lattes": lattes if lattes and lattes != "—" and lattes.lower() != "none" else "",
                "researchgate": rg if rg and rg != "—" and rg.lower() != "none" else "",
                "scholar": scholar if scholar and scholar != "—" and scholar.lower() != "none" else "",
                "areas": areas_list,
                "universities": uni_list if len(uni_list) > 1 else None,
                "university": uni_list[0] if len(uni_list) == 1 else None
            }
            colab_obj = {k: v for k, v in colab_obj.items() if v}
            collaborators.append(colab_obj)
        elif tipo == "Instituição Parceira":
            uni_obj = {
                "name": nome,
                "description": cargo,
                "img": caminho,
                "url": url_site
            }
            universities.append(uni_obj)
            
    print(f"Parcerias carregadas: {len(collaborators)} colaboradores, {len(universities)} instituições.")


# 3. SINCRONIZAR ARTIGOS (artigos.xlsx -> Papers.jsx)
artigos_file = os.path.join(DADOS_DIR, "artigos.xlsx")
if os.path.exists(artigos_file):
    wb = openpyxl.load_workbook(artigos_file, data_only=True)
    ws = wb.active
    
    all_papers = {}
    
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or len(row) < 3 or row[2] is None:
            continue
            
        row_id = row[2]
        try:
            pid = int(float(str(row_id).strip()))
        except (ValueError, TypeError):
            continue
            
        ano = str(row[3] or "").strip()
        titulo = str(row[4] or "").strip()
        autores = str(row[5] or "").strip()
        detalhes = str(row[6] or "").strip()
        detalhes = re.sub(r"\.\s*DOI\s*$", "", detalhes, flags=re.IGNORECASE).strip()
        detalhes = re.sub(r"\s+DOI\s*$", "", detalhes, flags=re.IGNORECASE).strip()
        doi = str(row[7] or "").strip()
        if doi and doi != "—" and not doi.startswith("http://") and not doi.startswith("https://"):
            if doi.startswith("10."):
                doi = f"https://doi.org/{doi}"
            else:
                doi = f"https://{doi}"
                
        tags = []
        for col_idx in range(9, 18):
            if col_idx < len(row) and row[col_idx]:
                t = str(row[col_idx]).strip()
                if t and t.lower() != "none" and t not in tags:
                    tags.append(t)
                    
        if not tags and len(row) > 8 and row[8]:
            raw_i = str(row[8]).strip()
            if raw_i and not raw_i.startswith("="):
                tags = [t.strip() for t in raw_i.split(",") if t.strip()]

        paper_obj = {
            "id": pid,
            "authors": autores,
            "title": titulo,
            "details": detalhes,
            "year": ano,
            "citations": "",
            "doi": doi,
            "tags": tags
        }
        
        if ano not in all_papers:
            all_papers[ano] = []
        all_papers[ano].append(paper_obj)
        
    total_papers = sum(len(v) for v in all_papers.values())
    print(f"Artigos carregados: {total_papers} artigos divididos em {len(all_papers)} anos.")


# 3.5 SINCRONIZAR PROJETOS (dados_site/projetos.xlsx -> src/data/projectsData.json)
projetos_file = os.path.join(DADOS_DIR, "projetos.xlsx")

if os.path.exists(projetos_file):
    wb = openpyxl.load_workbook(projetos_file)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        header_row = [str(cell or "").strip() for cell in rows[0]]
        projects_list = []
        
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            row_dict = {}
            for col_idx, col_name in enumerate(header_row):
                if col_name:
                    val = row[col_idx] if col_idx < len(row) else ""
                    row_dict[col_name] = str(val or "").strip()
            
            if row_dict.get("id", "").lower() in ["id", "id_projeto"]:
                continue
                
            projects_list.append(row_dict)
            
        json_path = os.path.join(BASE_DIR, "src", "data", "projectsData.json")
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(projects_list, f, ensure_ascii=False, indent=2)
        print(f"Projetos carregados: {len(projects_list)} projetos gravados em {json_path}.")

# 4. GRAVAR DADOS NOS COMPONENTES REACT
# A) Team.jsx
team_path = os.path.join(PAGES_DIR, "Team.jsx")
if os.path.exists(team_path) and professors:
    with open(team_path, "r", encoding="utf-8") as f:
        team_code = f.read()

    prof_json = json.dumps(professors, indent=4, ensure_ascii=False)
    res_json = json.dumps(researchers, indent=4, ensure_ascii=False)
    former_json = json.dumps(former_members, indent=4, ensure_ascii=False)

    team_code = re.sub(r"const professors = \[.*?\];", f"const professors = {prof_json};", team_code, flags=re.DOTALL)
    team_code = re.sub(r"const researchers = \[.*?\];", f"const researchers = {res_json};", team_code, flags=re.DOTALL)
    team_code = re.sub(r"const formerMembers = \[.*?\];", f"const formerMembers = {former_json};", team_code, flags=re.DOTALL)

    with open(team_path, "w", encoding="utf-8") as f:
        f.write(team_code)
    print("-> Team.jsx atualizado com sucesso!")

# B) Partners.jsx
partners_path = os.path.join(PAGES_DIR, "Partners.jsx")
if os.path.exists(partners_path) and universities:
    with open(partners_path, "r", encoding="utf-8") as f:
        partners_code = f.read()

    unis_json = json.dumps(universities, indent=4, ensure_ascii=False)
    colabs_json = json.dumps(collaborators, indent=4, ensure_ascii=False)

    partners_code = re.sub(r"const universities = \[.*?\];", f"const universities = {unis_json};", partners_code, flags=re.DOTALL)
    partners_code = re.sub(r"const collaborators = \[.*?\];", f"const collaborators = {colabs_json};", partners_code, flags=re.DOTALL)

    with open(partners_path, "w", encoding="utf-8") as f:
        f.write(partners_code)
    print("-> Partners.jsx atualizado com sucesso!")

# C) Papers.jsx
papers_path = os.path.join(PAGES_DIR, "Papers.jsx")
if os.path.exists(papers_path) and all_papers:
    with open(papers_path, "r", encoding="utf-8") as f:
        papers_code = f.read()

    papers_json = json.dumps(all_papers, indent=4, ensure_ascii=False)
    papers_code = re.sub(r"const allPapers = \{.*?\};\n\n  // Sort years", f"const allPapers = {papers_json};\n\n  // Sort years", papers_code, flags=re.DOTALL)

    with open(papers_path, "w", encoding="utf-8") as f:
        f.write(papers_code)
    print("-> Papers.jsx atualizado com sucesso!")

print("\nSincronizacao completa! O site foi atualizado com base nos arquivos .xlsx em dados_site/")
